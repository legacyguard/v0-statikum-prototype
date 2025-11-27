#!/usr/bin/env python3
"""
Script to generate mock ČSU agriculture time series Excel file.
"""
import csv
import sys

# Mock data for Czech agriculture time series
data = [
    ["Rok", "Produkce pšenice (tuny)", "Produkce ječmene (tuny)", "Produkce kukuřice (tuny)", "Počet hospodářství"],
    [2018, 5234567, 1876543, 876543, 26500],
    [2019, 5456789, 1923456, 912345, 26200],
    [2020, 5123456, 1845678, 845678, 25800],
    [2021, 5567890, 2012345, 967890, 25400],
    [2022, 5678901, 2098765, 1012345, 25100],
    [2023, 5789012, 2145678, 1098765, 24800],
    [2024, 5890123, 2187654, 1145678, 24500],
]

try:
    import openpyxl

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Časové řady - Zemědělství"

    # Write data
    for row in data:
        ws.append(row)

    # Format header
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Adjust column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 25

    # Save file
    output_path = sys.argv[1] if len(sys.argv) > 1 else "docs/csu-zemedelstvi-time-series.xlsx"
    wb.save(output_path)
    print(f"Excel file created: {output_path}")

except ImportError:
    print("openpyxl not installed, creating CSV instead...")
    # Fallback to CSV
    output_path = sys.argv[1] if len(sys.argv) > 1 else "docs/csu-zemedelstvi-time-series.csv"

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(data)

    print(f"CSV file created: {output_path}")
    print("Install openpyxl to create XLSX: pip3 install openpyxl")
